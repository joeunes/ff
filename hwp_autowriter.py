import os
import re
import sys
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. IFC 데이터 추출 및 유틸리티 모듈
# ---------------------------------------------------------------------------

def decode_ifc_string(s):
    if not s:
        return ""
    def replace_hex(match):
        hex_str = match.group(1)
        chars = []
        for i in range(0, len(hex_str), 4):
            hex_val = hex_str[i:i+4]
            if len(hex_val) == 4:
                chars.append(chr(int(hex_val, 16)))
        return "".join(chars)
    return re.sub(r'\\X2\\([0-9A-F]+)\\X0\\', replace_hex, s)

def calculate_mesh_volume(shape):
    verts = np.array(shape.geometry.verts).reshape(-1, 3)
    faces = np.array(shape.geometry.faces).reshape(-1, 3)
    total_volume = 0.0
    for face in faces:
        v0, v1, v2 = verts[face[0]], verts[face[1]], verts[face[2]]
        v = np.dot(v0, np.cross(v1, v2)) / 6.0
        total_volume += v
    return abs(total_volume)

def calculate_floor_area(shape):
    verts = np.array(shape.geometry.verts).reshape(-1, 3)
    faces = np.array(shape.geometry.faces).reshape(-1, 3)
    floor_area = 0.0
    for face in faces:
        v0, v1, v2 = verts[face[0]], verts[face[1]], verts[face[2]]
        cross_prod = np.cross(v1 - v0, v2 - v0)
        area = 0.5 * np.linalg.norm(cross_prod)
        norm = cross_prod / np.linalg.norm(cross_prod) if np.linalg.norm(cross_prod) > 0 else np.zeros(3)
        if abs(norm[2]) > 0.99:
            floor_area += area
    return (floor_area / 2.0)

def get_family_type_name(elem_name):
    """
    Revit에서 내보낸 객체 이름(예: 패밀리명:유형명:RevitID)에서 
    마지막의 고유 RevitID(숫자)를 제거하여 패밀리 유형 이름만 반환합니다.
    """
    if not elem_name:
        return "Unnamed"
    tokens = elem_name.split(":")
    if len(tokens) > 1 and tokens[-1].isdigit():
        return ":".join(tokens[:-1])
    return elem_name

# ---------------------------------------------------------------------------
# 2. 엑셀 워크북 빌더 및 스타일링 함수
# ---------------------------------------------------------------------------

def apply_auto_width(ws):
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1: # 타이틀 행 제외
                continue
            cell_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
            if cell_len > max_len:
                max_len = cell_len
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# ---------------------------------------------------------------------------
# 3. 데이터 파싱 및 엑셀 작성 메인 로직
# ---------------------------------------------------------------------------

def run_bim_to_excel(ifc_a_path, ifc_s_path, output_excel_path):
    import ifcopenshell
    import ifcopenshell.geom
    
    summary = {
        "arch_wall_area": 0.0,
        "arch_space_area": 0.0,
        "struct_element_qty": 0
    }
    arch_geom = []
    struct_geom = []
    
    settings = ifcopenshell.geom.settings()
    settings.set(settings.USE_WORLD_COORDS, True)
    
    wb = Workbook()
    default_sheet = wb.active
    
    # 공통 디자인 스타일 정의
    font_title = Font(name="맑은 고딕", size=15, bold=True, color="1E293B")
    font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="맑은 고딕", size=10, color="334155")
    font_total = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")
    
    fill_even_row = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="64748B")
    
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_total_row = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_double)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 층 정렬 기준 함수
    def storey_sort_key(name):
        if "지하" in name:
            match = re.search(r'\d+', name)
            return -int(match.group()) if match else -99
        elif "지상" in name:
            match = re.search(r'\d+', name)
            return int(match.group()) if match else 99
        elif "지붕" in name: return 100
        else: return 0

    # -----------------------------------------------------------------------
    # A. 건축 모델 (BA.ifc) 파싱 및 작성
    # -----------------------------------------------------------------------
    if os.path.exists(ifc_a_path):
        print(f"[건축 모델] {os.path.basename(ifc_a_path)} 로드 및 데이터 파싱 중...")
        model_a = ifcopenshell.open(ifc_a_path)
        
        # 층 정보 매핑 수집 (건축 모델 전체)
        spatial_relations_a = model_a.by_type("IfcRelContainedInSpatialStructure")
        element_to_storey_a = {}
        for rel in spatial_relations_a:
            parent = rel.RelatingStructure
            if parent.is_a("IfcBuildingStorey"):
                storey_name = decode_ifc_string(parent.Name)
                for elem in rel.RelatedElements:
                    element_to_storey_a[elem.id()] = storey_name
                    
        # 공간(룸) 이외에 IfcRelAggregates로 층에 포함되는 공간들 매핑
        aggregates_a = model_a.by_type("IfcRelAggregates")
        for rel in aggregates_a:
            parent = rel.RelatingObject
            if parent.is_a("IfcBuildingStorey"):
                storey_name = decode_ifc_string(parent.Name)
                for elem in rel.RelatedObjects:
                    element_to_storey_a[elem.id()] = storey_name
        
        # 1) 시트: 공간 일람표
        spaces = model_a.by_type("IfcSpace")
        space_list = []
        for s in spaces:
            level = element_to_storey_a.get(s.id(), "기타")
            num = decode_ifc_string(s.Name)
            name = decode_ifc_string(s.LongName)
            try:
                shape = ifcopenshell.geom.create_shape(settings, s)
                area = calculate_floor_area(shape)
            except Exception:
                area = 0.0
            space_list.append({"Level": level, "RoomNumber": num, "RoomName": name, "Area": area})
            
        df_spaces = pd.DataFrame(space_list)
        df_spaces = df_spaces.sort_values(by=["Level", "RoomNumber"]).reset_index(drop=True)
        summary["arch_space_area"] = float(df_spaces["Area"].sum())
        
        ws_spaces = wb.create_sheet(title="공간 일람표")
        ws_spaces["A1"].value = "BIM 공간 일람표 (건축 모델 기반)"
        ws_spaces["A1"].font = font_title
        ws_spaces.row_dimensions[1].height = 30
        
        fill_header_spaces = PatternFill(start_color="104F55", end_color="104F55", fill_type="solid") # Dark Teal
        headers_spaces = ["구분(층)", "실 번호", "실 이름", "면적 (㎡)"]
        for col_idx, h in enumerate(headers_spaces, 1):
            cell = ws_spaces.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_spaces
            cell.alignment = align_center
            cell.border = border_cell
        ws_spaces.row_dimensions[3].height = 25
        
        start_row = 4
        for idx, (row_idx, row) in enumerate(df_spaces.iterrows()):
            r = start_row + idx
            c1 = ws_spaces.cell(row=r, column=1, value=row["Level"])
            c2 = ws_spaces.cell(row=r, column=2, value=row["RoomNumber"])
            c3 = ws_spaces.cell(row=r, column=3, value=row["RoomName"])
            c4 = ws_spaces.cell(row=r, column=4, value=row["Area"])
            c1.alignment = align_center
            c2.alignment = align_center
            c3.alignment = align_left
            c4.alignment = align_right
            c4.number_format = "#,##0.00"
            for cell in [c1, c2, c3, c4]:
                cell.font = font_body
                cell.border = border_cell
                if idx % 2 == 1: cell.fill = fill_even_row
            ws_spaces.row_dimensions[r].height = 18
            
        total_row = start_row + len(df_spaces)
        t1 = ws_spaces.cell(row=total_row, column=1, value="총합계")
        t1.font = font_total
        t1.alignment = align_center
        t1.fill = fill_total
        t1.border = border_total_row
        ws_spaces.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        for col_idx in range(1, 4):
            ws_spaces.cell(row=total_row, column=col_idx).border = border_total_row
            ws_spaces.cell(row=total_row, column=col_idx).fill = fill_total
        t4 = ws_spaces.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
        t4.font = font_total
        t4.fill = fill_total
        t4.alignment = align_right
        t4.border = border_total_row
        t4.number_format = "#,##0.00"
        ws_spaces.row_dimensions[total_row].height = 24
        apply_auto_width(ws_spaces)
        print(f"  [완료] 공간 일람표 시트 작성 완료 ({len(df_spaces)}개 실)")

        # 2) 건축부재들 추출 및 시트 작성 (벽체, 슬래브, 문, 창문, 마감재)
        arch_categories = {
            "IfcWall": ("건축_벽체", "4A154B"),       # Plum
            "IfcSlab": ("건축_슬래브", "6B114D"),     # Deep Pink
            "IfcDoor": ("건축_문", "3F0E40"),         # Eggplant
            "IfcWindow": ("건축_창문", "1F5B52"),     # Ocean Teal
            "IfcCovering": ("건축_마감재", "7C3AED")   # Purple
        }
        
        for ifc_class, (sheet_title, theme_color) in arch_categories.items():
            elements = model_a.by_type(ifc_class)
            print(f"  {ifc_class} 건축부재 계산 중 ({len(elements)}개)...")
            
            is_wall = (ifc_class == "IfcWall")
            qty_label = "면적" if is_wall else "체적"
            col_header_val = "면적 합계 (㎡)" if is_wall else "체적 합계 (㎥)"
            
            cat_list = []
            for elem in elements:
                try:
                    shape = ifcopenshell.geom.create_shape(settings, elem)
                    vol = calculate_mesh_volume(shape)
                    storey = element_to_storey_a.get(elem.id(), "기타")
                    raw_name = elem.Name if elem.Name else "Unnamed"
                    family_type = get_family_type_name(raw_name)
                    
                    if is_wall:
                        # bounding box로 두께 계산
                        geom = getattr(shape, "geometry")
                        verts = np.array(geom.verts).reshape(-1, 3)
                        min_coords = np.min(verts, axis=0)
                        max_coords = np.max(verts, axis=0)
                        bbox = max_coords - min_coords
                        thickness = min(bbox[0], bbox[1])
                        qty_val = (vol / thickness) if thickness > 0 else 0.0
                    else:
                        qty_val = vol
                    
                    cat_list.append({
                        "층(Level)": storey,
                        "패밀리유형": family_type,
                        qty_label: qty_val
                    })
                    
                    # Collect geometry for 3D viewer
                    geom = getattr(shape, "geometry")
                    arch_geom.append({
                        "id": elem.id(),
                        "class": ifc_class,
                        "name": decode_ifc_string(elem.Name) if elem.Name else "Unnamed",
                        "type": family_type,
                        "storey": storey,
                        "quantity": f"{qty_val:.2f} m²" if is_wall else f"{qty_val:.2f} m³",
                        "verts": list(geom.verts),
                        "faces": list(geom.faces)
                    })
                except Exception:
                    pass
            
            df_cat = pd.DataFrame(cat_list)
            if df_cat.empty:
                continue
                
            # 유형별, 층별 그룹 합산
            df_grouped = df_cat.groupby(["층(Level)", "패밀리유형"]).agg(
                수량합계=(qty_label, "sum"),
                수량EA=(qty_label, "count") # 개체수
            ).reset_index()
            
            df_grouped["SortKey"] = df_grouped["층(Level)"].map(storey_sort_key)
            df_grouped = df_grouped.sort_values(by=["SortKey", "패밀리유형"]).drop(columns=["SortKey"]).reset_index(drop=True)
            if is_wall:
                summary["arch_wall_area"] = float(df_grouped["수량합계"].sum())
            
            # 시트 생성
            ws_cat = wb.create_sheet(title=sheet_title)
            ws_cat["A1"].value = f"BIM 건축분야 {sheet_title.split('_')[1]} 수량산출서 (유형별 집계)"
            ws_cat["A1"].font = font_title
            ws_cat.row_dimensions[1].height = 30
            
            fill_header_cat = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
            headers_cat = ["구분(층)", "패밀리 및 유형 이름", "수량 (EA)", col_header_val]
            
            for col_idx, h in enumerate(headers_cat, 1):
                cell = ws_cat.cell(row=3, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header_cat
                cell.alignment = align_center
                cell.border = border_cell
            ws_cat.row_dimensions[3].height = 25
            
            # 데이터 쓰기
            start_row = 4
            for idx, (row_idx, row) in enumerate(df_grouped.iterrows()):
                r = start_row + idx
                
                c1 = ws_cat.cell(row=r, column=1, value=row["층(Level)"])
                c2 = ws_cat.cell(row=r, column=2, value=row["패밀리유형"])
                c3 = ws_cat.cell(row=r, column=3, value=row["수량EA"])
                c4 = ws_cat.cell(row=r, column=4, value=row["수량합계"])
                
                c1.alignment = align_center
                c2.alignment = align_left
                c3.alignment = align_right
                c4.alignment = align_right
                
                c3.number_format = "#,##0"
                c4.number_format = "#,##0.00"
                
                for cell in [c1, c2, c3, c4]:
                    cell.font = font_body
                    cell.border = border_cell
                    if idx % 2 == 1:
                        cell.fill = fill_even_row
                ws_cat.row_dimensions[r].height = 18
                
            # 합계 행
            total_row = start_row + len(df_grouped)
            t1 = ws_cat.cell(row=total_row, column=1, value="총합계")
            t1.font = font_total
            t1.alignment = align_center
            t1.fill = fill_total
            t1.border = border_total_row
            
            ws_cat.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
            for col_idx in range(1, 3):
                ws_cat.cell(row=total_row, column=col_idx).border = border_total_row
                ws_cat.cell(row=total_row, column=col_idx).fill = fill_total
                
            t3 = ws_cat.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})")
            t4 = ws_cat.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
            
            for cell in [t3, t4]:
                cell.font = font_total
                cell.fill = fill_total
                cell.alignment = align_right
                cell.border = border_total_row
            t3.number_format = "#,##0"
            t4.number_format = "#,##0.00"
            ws_cat.row_dimensions[total_row].height = 24
            
            apply_auto_width(ws_cat)
            print(f"  [완료] {sheet_title} 시트 작성 완료 ({len(df_grouped)}개 유형 집계)")

    else:
        print(f"[경고] 건축 모델 {ifc_a_path} 파일이 존재하지 않아 건축 시트 작성을 건너뜁니다.")

    # -----------------------------------------------------------------------
    # B. 구조 모델 (BS.ifc) 파싱 및 작성
    # -----------------------------------------------------------------------
    df_detail = pd.DataFrame()
    if os.path.exists(ifc_s_path):
        print(f"[구조 모델] {os.path.basename(ifc_s_path)} 로드 및 수량 정보 추출 중...")
        model_s = ifcopenshell.open(ifc_s_path)
        
        spatial_relations_s = model_s.by_type("IfcRelContainedInSpatialStructure")
        element_to_storey_s = {}
        for rel in spatial_relations_s:
            parent = rel.RelatingStructure
            if parent.is_a("IfcBuildingStorey"):
                storey_name = decode_ifc_string(parent.Name)
                for elem in rel.RelatedElements:
                    element_to_storey_s[elem.id()] = storey_name
                    
        types = {"IfcColumn": "기둥", "IfcBeam": "보", "IfcWall": "벽체"}
        detailed_list = []
        
        for t_ifc, t_name in types.items():
            elements = model_s.by_type(t_ifc)
            print(f"  {t_name} 부재 체적 계산 중 ({len(elements)}개)...")
            for elem in elements:
                try:
                    shape = ifcopenshell.geom.create_shape(settings, elem)
                    vol = calculate_mesh_volume(shape)
                    storey = element_to_storey_s.get(elem.id(), "기타")
                    raw_name = elem.Name if elem.Name else "Unnamed"
                    family_type = get_family_type_name(raw_name)
                    
                    detailed_list.append({
                        "ID": elem.id(),
                        "층(Level)": storey,
                        "부재종류": t_name,
                        "패밀리유형": family_type,
                        "체적(㎥)": vol
                    })
                    
                    # Collect geometry for 3D viewer
                    geom = getattr(shape, "geometry")
                    struct_geom.append({
                        "id": elem.id(),
                        "class": t_ifc,
                        "name": decode_ifc_string(elem.Name) if elem.Name else "Unnamed",
                        "type": family_type,
                        "storey": storey,
                        "quantity": f"{vol:.2f} m³",
                        "verts": list(geom.verts),
                        "faces": list(geom.faces)
                    })
                except Exception:
                    pass
                    
        df_detail = pd.DataFrame(detailed_list)
        if not df_detail.empty:
            summary["struct_element_qty"] = len(df_detail)
        
        if not df_detail.empty:
            # 1) 구조수량 요약
            pivot_df = df_detail.pivot_table(
                index="층(Level)", 
                columns="부재종류", 
                values="체적(㎥)", 
                aggfunc="sum", 
                fill_value=0.0
            ).reset_index()
            
            for col in ["기둥", "보", "벽체"]:
                if col not in pivot_df.columns:
                    pivot_df[col] = 0.0
                    
            pivot_df["SortKey"] = pivot_df["층(Level)"].map(storey_sort_key)
            pivot_df = pivot_df.sort_values(by="SortKey").drop(columns=["SortKey"]).reset_index(drop=True)
            
            ws_struct_summary = wb.create_sheet(title="구조수량 요약")
            ws_struct_summary["A1"].value = "BIM 구조분야 콘크리트 수량산출 요약서"
            ws_struct_summary["A1"].font = font_title
            ws_struct_summary.row_dimensions[1].height = 30
            
            fill_header_struct = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy
            headers_struct = ["구분(층)", "기둥 (㎥)", "보 (㎥)", "벽체 (㎥)", "합계 (㎥)"]
            
            for col_idx, h in enumerate(headers_struct, 1):
                cell = ws_struct_summary.cell(row=3, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header_struct
                cell.alignment = align_center
                cell.border = border_cell
            ws_struct_summary.row_dimensions[3].height = 25
            
            start_row = 4
            for idx, (row_idx, row) in enumerate(pivot_df.iterrows()):
                r = start_row + idx
                
                c1 = ws_struct_summary.cell(row=r, column=1, value=row["층(Level)"])
                c2 = ws_struct_summary.cell(row=r, column=2, value=row["기둥"])
                c3 = ws_struct_summary.cell(row=r, column=3, value=row["보"])
                c4 = ws_struct_summary.cell(row=r, column=4, value=row["벽체"])
                c5 = ws_struct_summary.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
                
                c1.alignment = align_center
                for cell in [c2, c3, c4, c5]:
                    cell.alignment = align_right
                    cell.number_format = "#,##0.00"
                    
                for cell in [c1, c2, c3, c4, c5]:
                    cell.font = font_body
                    cell.border = border_cell
                    if idx % 2 == 1:
                        cell.fill = fill_even_row
                ws_struct_summary.row_dimensions[r].height = 20
                
            # 합계 행
            total_row = start_row + len(pivot_df)
            t1 = ws_struct_summary.cell(row=total_row, column=1, value="총합계")
            t1.font = font_total
            t1.alignment = align_center
            t1.fill = fill_total
            t1.border = border_total_row
            
            t2 = ws_struct_summary.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})")
            t3 = ws_struct_summary.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})")
            t4 = ws_struct_summary.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
            t5 = ws_struct_summary.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})")
            
            for cell in [t2, t3, t4, t5]:
                cell.font = font_total
                cell.fill = fill_total
                cell.alignment = align_right
                cell.border = border_total_row
                cell.number_format = "#,##0.00"
            ws_struct_summary.row_dimensions[total_row].height = 24
            
            apply_auto_width(ws_struct_summary)
            print("  [완료] 구조수량 요약 시트 작성 완료")
            
            # 2) 구조 부재별 개별 시트 (구조_기둥, 구조_보, 구조_벽체)
            categories = {
                "기둥": ("구조_기둥", "1F2A4A"), 
                "보": ("구조_보", "2E5A88"),    
                "벽체": ("구조_벽체", "475569")  
            }
            
            for cat_name, (sheet_title, theme_color) in categories.items():
                df_cat = df_detail[df_detail["부재종류"] == cat_name].copy()
                if df_cat.empty:
                    continue
                
                df_grouped = df_cat.groupby(["층(Level)", "패밀리유형"]).agg(
                    체적합계=("체적(㎥)", "sum"),
                    수량EA=("ID", "count")
                ).reset_index()
                
                df_grouped["SortKey"] = df_grouped["층(Level)"].map(storey_sort_key)
                df_grouped = df_grouped.sort_values(by=["SortKey", "패밀리유형"]).drop(columns=["SortKey"]).reset_index(drop=True)
                
                ws_cat = wb.create_sheet(title=sheet_title)
                ws_cat["A1"].value = f"BIM 구조분야 {cat_name} 수량산출서 (유형별 집계)"
                ws_cat["A1"].font = font_title
                ws_cat.row_dimensions[1].height = 30
                
                fill_header_cat = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
                headers_cat = ["구분(층)", "패밀리 및 유형 이름", "수량 (EA)", "체적 합계 (㎥)"]
                
                for col_idx, h in enumerate(headers_cat, 1):
                    cell = ws_cat.cell(row=3, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_header_cat
                    cell.alignment = align_center
                    cell.border = border_cell
                ws_cat.row_dimensions[3].height = 25
                
                start_row = 4
                for idx, (row_idx, row) in enumerate(df_grouped.iterrows()):
                    r = start_row + idx
                    
                    c1 = ws_cat.cell(row=r, column=1, value=row["층(Level)"])
                    c2 = ws_cat.cell(row=r, column=2, value=row["패밀리유형"])
                    c3 = ws_cat.cell(row=r, column=3, value=row["수량EA"])
                    c4 = ws_cat.cell(row=r, column=4, value=row["체적합계"])
                    
                    c1.alignment = align_center
                    c2.alignment = align_left
                    c3.alignment = align_right
                    c4.alignment = align_right
                    
                    c3.number_format = "#,##0"
                    c4.number_format = "#,##0.00"
                    
                    for cell in [c1, c2, c3, c4]:
                        cell.font = font_body
                        cell.border = border_cell
                        if idx % 2 == 1:
                            cell.fill = fill_even_row
                    ws_cat.row_dimensions[r].height = 18
                    
                # 합계 행
                total_row = start_row + len(df_grouped)
                t1 = ws_cat.cell(row=total_row, column=1, value="총합계")
                t1.font = font_total
                t1.alignment = align_center
                t1.fill = fill_total
                t1.border = border_total_row
                
                ws_cat.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
                for col_idx in range(1, 3):
                    ws_cat.cell(row=total_row, column=col_idx).border = border_total_row
                    ws_cat.cell(row=total_row, column=col_idx).fill = fill_total
                    
                t3 = ws_cat.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})")
                t4 = ws_cat.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
                
                for cell in [t3, t4]:
                    cell.font = font_total
                    cell.fill = fill_total
                    cell.alignment = align_right
                    cell.border = border_total_row
                t3.number_format = "#,##0"
                t4.number_format = "#,##0.00"
                ws_cat.row_dimensions[total_row].height = 24
                
                apply_auto_width(ws_cat)
                print(f"  [완료] {sheet_title} 시트 작성 완료 ({len(df_grouped)}개 유형 집계)")
                
    else:
        print(f"[경고] 구조 모델 {ifc_s_path} 파일이 존재하지 않아 구조 수량 시트 작성을 건너뜁니다.")

    # -----------------------------------------------------------------------
    # C. 워크북 정리 및 저장
    # -----------------------------------------------------------------------
    if "Sheet" in wb.sheetnames:
        wb.remove(default_sheet)
        
    import time
    saved_path = output_excel_path
    try:
        wb.save(output_excel_path)
        print(f"\n[성공] 최종 통합 수량산출서 엑셀 저장 완료: {output_excel_path}")
    except PermissionError:
        base, ext = os.path.splitext(output_excel_path)
        saved = False
        for suffix in ["_new", f"_new_{int(time.time())}", "_v2", "_v3"]:
            alt_path = f"{base}{suffix}{ext}"
            try:
                wb.save(alt_path)
                print(f"\n[경고] {output_excel_path} 파일이 이미 열려 있어 덮어쓸 수 없습니다.")
                print(f"       대신 다른 이름으로 저장합니다: {alt_path}")
                saved_path = alt_path
                saved = True
                break
            except PermissionError:
                continue
        if not saved:
            raise PermissionError("엑셀 파일을 저장할 수 없습니다. 모든 후보 경로가 잠겨 있습니다.")
            
    # Save geometry JSON files
    if os.path.exists(ifc_a_path) and arch_geom:
        arch_json_path = saved_path.replace(".xlsx", "_arch.json")
        with open(arch_json_path, "w") as f:
            json.dump(arch_geom, f)
            
    if os.path.exists(ifc_s_path) and struct_geom:
        struct_json_path = saved_path.replace(".xlsx", "_struct.json")
        with open(struct_json_path, "w") as f:
            json.dump(struct_geom, f)
            
    summary["saved_excel_path"] = saved_path
    return summary

def main():
    base_dir = os.path.abspath(os.path.dirname(__file__))
    ifc_a_path = os.path.join(base_dir, "BA.ifc")
    ifc_s_path = os.path.join(base_dir, "BS.ifc")
    
    excel_output = os.path.join(base_dir, "BIM_수량_및_공간_산출서.xlsx")
    
    print("=== BIM 데이터(공간 및 수량) 통합 엑셀 보고서 작성 프로그램 ===")
    run_bim_to_excel(ifc_a_path, ifc_s_path, excel_output)

if __name__ == '__main__':
    main()
